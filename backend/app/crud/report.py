import uuid
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.models.sale import Sale, SaleLineItem, SaleStatus
from app.models.product import FrameProduct
from app.models.customer import Customer


# ------------------------------------------------------------------ #
# Sales list / summary
# ------------------------------------------------------------------ #

def get_sales_list(
    db: Session, branch_id: uuid.UUID,
    start: date, end: date,
    limit: int = 50, offset: int = 0,
    status: str | None = None,
):
    q = (
        db.query(Sale)
        .filter(Sale.branch_id == branch_id)
        .filter(Sale.created_at >= datetime.combine(start, datetime.min.time()))
        .filter(Sale.created_at < datetime.combine(end + timedelta(days=1), datetime.min.time()))
    )
    if status:
        q = q.filter(Sale.status == status)
    return q.order_by(Sale.created_at.desc()).offset(offset).limit(limit).all()


def get_all_branches_sales_list(
    db: Session, start: date, end: date,
    branch_id: uuid.UUID | None = None,
    limit: int = 200, offset: int = 0,
):
    q = (
        db.query(Sale)
        .filter(Sale.created_at >= datetime.combine(start, datetime.min.time()))
        .filter(Sale.created_at < datetime.combine(end + timedelta(days=1), datetime.min.time()))
        .filter(Sale.status == SaleStatus.ACTIVE)
    )
    if branch_id:
        q = q.filter(Sale.branch_id == branch_id)
    return q.order_by(Sale.created_at.desc()).offset(offset).limit(limit).all()


def get_sales_summary(db: Session, branch_id: uuid.UUID, period: str) -> dict:
    now = datetime.utcnow()
    if period == "daily":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "weekly":
        start = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "monthly":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    elif period == "yearly":
        start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        raise ValueError("period must be daily/weekly/monthly/yearly")

    result = (
        db.query(
            func.coalesce(func.sum(Sale.total), 0).label("total_revenue"),
            func.count(Sale.id).label("total_sales_count"),
            func.coalesce(func.sum(Sale.discount), 0).label("total_discount"),
        )
        .filter(Sale.branch_id == branch_id, Sale.created_at >= start, Sale.status == SaleStatus.ACTIVE)
        .first()
    )
    return {
        "branch_id": str(branch_id),
        "period": period,
        "period_start": start.date(),
        "period_end": now.date(),
        "total_revenue": result.total_revenue,
        "total_sales_count": result.total_sales_count,
        "total_discount": result.total_discount,
    }


def get_all_branches_summary(db: Session, period: str) -> dict:
    now = datetime.utcnow()
    if period == "daily":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "weekly":
        start = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "monthly":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    elif period == "yearly":
        start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        raise ValueError("period must be daily/weekly/monthly/yearly")

    from app.models.branch import Branch

    rows = (
        db.query(
            Sale.branch_id,
            func.coalesce(func.sum(Sale.total), 0).label("revenue"),
            func.count(Sale.id).label("sales_count"),
            func.coalesce(func.sum(Sale.discount), 0).label("discount"),
        )
        .filter(Sale.created_at >= start, Sale.status == SaleStatus.ACTIVE)
        .group_by(Sale.branch_id)
        .all()
    )

    # Resolve branch names in a single extra query — avoids JOIN grouping issues
    branch_ids = [r.branch_id for r in rows]
    branches   = db.query(Branch).filter(Branch.id.in_(branch_ids)).all() if branch_ids else []
    branch_map = {b.id: b for b in branches}

    return {
        "period": period,
        "period_start": start.date(),
        "period_end": now.date(),
        "branches": [
            {
                "branch_id":   str(r.branch_id),
                "branch_name": branch_map[r.branch_id].name if r.branch_id in branch_map else str(r.branch_id)[:8],
                "branch_code": branch_map[r.branch_id].code if r.branch_id in branch_map else str(r.branch_id)[:8],
                "revenue":     r.revenue,
                "sales_count": r.sales_count,
                "discount":    r.discount,
            }
            for r in rows
        ],
        "total_revenue":     sum(r.revenue for r in rows),
        "total_sales_count": sum(r.sales_count for r in rows),
    }


def get_eod_summary(db: Session, branch_id: uuid.UUID, target_date: date) -> dict:
    start = datetime.combine(target_date, datetime.min.time())
    end = datetime.combine(target_date + timedelta(days=1), datetime.min.time())
    result = (
        db.query(
            func.coalesce(func.sum(Sale.total), 0).label("total"),
            func.count(Sale.id).label("count"),
            func.coalesce(func.sum(Sale.discount), 0).label("discount"),
            func.coalesce(func.sum(Sale.cash_tendered), 0).label("cash_in"),
        )
        .filter(
            Sale.branch_id == branch_id,
            Sale.created_at >= start,
            Sale.created_at < end,
            Sale.status == SaleStatus.ACTIVE,
        )
        .first()
    )
    return {
        "branch_id": str(branch_id),
        "date": target_date,
        "total_sales": result.count,
        "total_revenue": result.total,
        "total_discount": result.discount,
        "cash_collected": result.cash_in,
    }


def get_customer_purchase_history(db: Session, customer_id: uuid.UUID):
    return (
        db.query(Sale)
        .filter(Sale.customer_id == customer_id, Sale.status == SaleStatus.ACTIVE)
        .order_by(Sale.created_at.desc())
        .all()
    )


# ------------------------------------------------------------------ #
# Excel export (openpyxl)
# ------------------------------------------------------------------ #

def build_excel_report(
    db: Session,
    start: date, end: date,
    branch_id: uuid.UUID | None = None,
) -> BytesIO:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(fill_type="solid", fgColor="1E3A5F")

    def style_header(ws, headers: list[str]):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

    def auto_width(ws):
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=0)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 50)

    # Build lookup maps for branches and customers once
    from app.models.branch import Branch
    all_branches = db.query(Branch).all()
    branch_name_map: dict = {b.id: b.name for b in all_branches}

    all_customers = db.query(Customer).all()
    customer_phone_map: dict = {c.id: (c.phone or c.full_name or str(c.id)[:8]) for c in all_customers}

    # ---- Sheet 1: Sales Detail ----
    ws1 = wb.active
    ws1.title = "Sales Detail"
    headers1 = [
        "Invoice #", "Branch", "Date", "Customer Phone",
        "Items Bought", "Discount Type", "Discount %",
        "Subtotal", "Tax", "Total", "Payment Method", "Status",
    ]
    style_header(ws1, headers1)

    sales = get_all_branches_sales_list(db, start, end, branch_id=branch_id, limit=10000)
    for sale in sales:
        # Sum all line item quantities for "items bought"
        items_bought = sum(li.quantity for li in sale.line_items) if sale.line_items else 0
        # Resolve branch name
        branch_name = branch_name_map.get(sale.branch_id, str(sale.branch_id)[:8])
        # Resolve customer phone (fall back to name, then blank)
        customer_str = customer_phone_map.get(sale.customer_id, "") if sale.customer_id else ""

        ws1.append([
            sale.invoice_number,
            branch_name,
            sale.created_at.strftime("%Y-%m-%d %H:%M") if sale.created_at else "",
            customer_str,
            items_bought,
            sale.discount_type.value if sale.discount_type else "",
            float(sale.discount_pct or 0),
            float(sale.subtotal or 0),
            float(sale.tax_amount or 0),
            float(sale.total or 0),
            sale.payment_method.value if sale.payment_method else "",
            sale.status.value if sale.status else "",
        ])
    auto_width(ws1)

    # ---- Sheet 2: Inventory Snapshot ----
    ws2 = wb.create_sheet("Inventory Snapshot")
    headers2 = ["Branch", "SKU", "Barcode", "Name", "Brand", "Category",
                "Qty", "Reorder Threshold", "Selling Price", "Cost Price", "Status"]
    style_header(ws2, headers2)

    inv_q = db.query(FrameProduct)
    if branch_id:
        inv_q = inv_q.filter(FrameProduct.branch_id == branch_id)
    for fp in inv_q.all():
        ws2.append([
            branch_name_map.get(fp.branch_id, str(fp.branch_id)[:8]),
            fp.sku or fp.product_code,
            fp.barcode,
            fp.name,
            fp.brand or "",
            fp.category or "",
            fp.quantity,
            fp.reorder_threshold or 0,
            float(fp.selling_price),
            float(fp.cost_price or 0),
            "Active" if fp.is_active else "Inactive",
        ])
    auto_width(ws2)

    # ---- Sheet 3: Low Stock ----
    ws3 = wb.create_sheet("Low Stock")
    headers3 = ["Branch", "Barcode", "Name", "Qty", "Reorder Threshold", "Shortfall"]
    style_header(ws3, headers3)

    ls_q = db.query(FrameProduct).filter(
        FrameProduct.is_active == True,  # noqa: E712
        FrameProduct.quantity <= FrameProduct.reorder_threshold,
    )
    if branch_id:
        ls_q = ls_q.filter(FrameProduct.branch_id == branch_id)
    for fp in ls_q.all():
        ws3.append([
            branch_name_map.get(fp.branch_id, str(fp.branch_id)[:8]),
            fp.barcode, fp.name,
            fp.quantity, fp.reorder_threshold or 0,
            (fp.reorder_threshold or 0) - fp.quantity,
        ])
    auto_width(ws3)

    # ---- Sheet 4: Discount Usage ----
    ws4 = wb.create_sheet("Discount Usage")
    headers4 = ["Discount Type", "# Sales", "Total Discount Amount"]
    style_header(ws4, headers4)

    disc_q = (
        db.query(
            Sale.discount_type,
            func.count(Sale.id).label("cnt"),
            func.sum(Sale.discount).label("total_disc"),
        )
        .filter(
            Sale.created_at >= datetime.combine(start, datetime.min.time()),
            Sale.created_at < datetime.combine(end + timedelta(days=1), datetime.min.time()),
            Sale.status == SaleStatus.ACTIVE,
        )
        .group_by(Sale.discount_type)
    )
    if branch_id:
        disc_q = disc_q.filter(Sale.branch_id == branch_id)
    for row in disc_q.all():
        ws4.append([
            row.discount_type.value if row.discount_type else "none",
            row.cnt,
            float(row.total_disc or 0),
        ])
    auto_width(ws4)

    # ---- Sheet 5: Membership Summary ----
    ws5 = wb.create_sheet("Membership Summary")
    headers5 = ["Customer Name", "Phone", "Purchase Count", "Loyalty Points",
                "Discount Level", "Tier"]
    style_header(ws5, headers5)

    from app.services.discount import get_membership_tier
    for c in all_customers:
        if not c.is_active:
            continue
        _, tier_name = get_membership_tier(db, c.purchase_count)
        ws5.append([
            c.full_name, c.phone or "", c.purchase_count,
            c.loyalty_points, c.discount_level,
            tier_name or "—",
        ])
    auto_width(ws5)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
